import os
import shutil
import threading
from datetime import datetime
from time import sleep
from typing import Any, Dict, List, Optional

import flet as ft
import pandas as pd
import numpy as np
import sympy as sp
import io
import base64
import matplotlib
matplotlib.use("Agg")
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

import sys
try:
    # Try import from installed path first
    from dupr_client import DuprClient  # type: ignore
    _IMPORT_ERROR = None
except Exception as import_error:  # pragma: no cover
    # Fallback: look for local cloned repo folder 'duprly' next to this app
    _IMPORT_ERROR = None
    try:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        duprly_dir = os.path.join(current_dir, "duprly")
        if os.path.isdir(duprly_dir):
            sys.path.append(duprly_dir)
            from dupr_client import DuprClient  # type: ignore
        else:
            _IMPORT_ERROR = import_error
    except Exception as import_error2:
        _IMPORT_ERROR = import_error2


def get_club_data(
    client: "DuprClient",
    club_id: str,
    max_members: int = 20,
    max_matches: int = 5,
    matches_per_player: int = 10,
    log: Optional[callable] = None,
    stop_event: Optional[threading.Event] = None,
) -> Dict[str, Any]:
    def _log(message: str) -> None:
        if log:
            log(message)

    _log(f"Bắt đầu crawl dữ liệu cho club ID: {club_id}")

    # 1. Members
    _log("Đang lấy danh sách thành viên…")
    status, members = client.get_members_by_club(club_id)
    if not status:
        raise RuntimeError("Không lấy được danh sách thành viên")

    members = members[: max_members or 0]

    # 2. Player profiles
    _log("\nĐang lấy thông tin chi tiết thành viên…")
    player_profiles: List[Dict[str, Any]] = []
    for idx, member in enumerate(members):
        if stop_event is not None and stop_event.is_set():
            _log("Đã dừng lấy profile theo yêu cầu người dùng.")
            break
        player_id = str(member.get("id"))
        try:
            _log(f"  Đang xử lý player {idx + 1}/{len(members)}: {player_id}")

            profile: Dict[str, Any] = {
                "id": player_id,
                "duprId": member.get("duprId"),
                "fullName": member.get("fullName"),
                "gender": member.get("gender"),
                "age": member.get("age"),
                "shortAddress": member.get("shortAddress"),
                "singles": member.get("singles"),
                "doubles": member.get("doubles"),
                "singlesVerified": member.get("singlesVerified"),
                "doublesVerified": member.get("doublesVerified"),
                "singlesReliability": member.get("singlesReliability"),
                "doublesReliability": member.get("doublesReliability"),
            }

            status, detailed_profile = client.get_player(player_id)
            if detailed_profile:
                profile.update(
                    {
                        "singles_rating": detailed_profile.get("singles", {}).get(
                            "display", "N/A"
                        ),
                        "doubles_rating": detailed_profile.get("doubles", {}).get(
                            "display", "N/A"
                        ),
                    }
                )

            player_profiles.append(profile)
            sleep(1)
        except Exception as exc:  # pragma: no cover - best effort logging
            _log(f"Lỗi khi lấy profile {player_id}: {exc}")

    # 3. Match history
    _log("\nĐang lấy lịch sử trận đấu…")
    match_history: Dict[str, List[Dict[str, Any]]] = {}
    player_ids = [str(m["id"]) for m in members[: max_matches or 0]]

    for idx, pid in enumerate(player_ids):
        if stop_event is not None and stop_event.is_set():
            _log("Đã dừng lấy lịch sử trận đấu theo yêu cầu người dùng.")
            break
        _log(f"\nĐang lấy lịch sử trận đấu cho player {idx + 1}/{len(player_ids)}: {pid}")
        offset: Optional[int] = 0
        match_list: List[Dict[str, Any]] = []
        retry_count = 0
        max_retries = 3

        while offset is not None and retry_count < max_retries:
            if stop_event is not None and stop_event.is_set():
                _log("Dừng ở giữa khi lấy lịch sử trận đấu…")
                break
            try:
                page_data = {
                    "filters": {},
                    "sort": {"order": "DESC", "parameter": "MATCH_DATE"},
                    "limit": matches_per_player,
                    "offset": offset,
                }

                r = client.dupr_post(
                    f"/player/v1.0/{pid}/history",
                    json_data=page_data,
                    name="get_member_match_history",
                )

                if r.status_code != 200:
                    _log(f"Lỗi khi lấy lịch sử trận đấu (status: {r.status_code})")
                    if r.status_code == 403:
                        _log("Thử refresh token…")
                        client.login_user(client.email, client.password)  # type: ignore[attr-defined]
                        retry_count += 1
                        continue
                    break

                data = r.json()
                offset, hits = client.handle_paging(data)

                if not hits:
                    _log("Không có thêm trận đấu nào")
                    break

                _log(f"Đã lấy được {len(hits)} trận đấu")
                match_list.extend(hits)
                sleep(1)

            except Exception as exc:  # pragma: no cover - best effort logging
                _log(f"Lỗi khi xử lý: {exc}")
                retry_count += 1
                sleep(2)
                continue

        if match_list:
            match_history[pid] = match_list
        else:
            _log(f" Không lấy được lịch sử trận đấu cho player {pid}")

    return {
        "club_info": {
            "id": club_id,
            "name": members[0].get("clubName") if members else "",
            "total_members": len(members),
            "scraped_matches": sum(len(m) for m in match_history.values()),
        },
        "members": members,
        "player_profiles": player_profiles,
        "match_history": match_history,
    }


def export_to_excel(data: Dict[str, Any], filename_prefix: str = "dupr_export") -> str:
    wb = Workbook()

    # Sheet 1: Club Info
    ws_info = wb.active
    ws_info.title = "Club Info"

    club_info = [
        ["Club ID", data["club_info"]["id"]],
        ["Club Name", data["club_info"].get("name", "N/A")],
        ["Total Members", data["club_info"]["total_members"]],
        ["Players with Match History", len(data["match_history"])],
        ["Total Matches Scraped", data["club_info"]["scraped_matches"]],
        ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    for row in club_info:
        ws_info.append(row)

    # Sheet 2: Members
    if data["members"]:
        df_members = pd.DataFrame(data["members"]) if data["members"] else pd.DataFrame()
        member_cols = {
            "id": "Player ID",
            "duprId": "DUPR ID",
            "fullName": "Full Name",
            "gender": "Gender",
            "age": "Age",
            "singles": "Singles Rating",
            "doubles": "Doubles Rating",
            "shortAddress": "Location",
            "email": "Email",
            "phoneNumber": "Phone",
        }
        available_cols = [c for c in member_cols.keys() if c in df_members.columns]
        df_members = df_members[available_cols].rename(columns=member_cols)

        ws_members = wb.create_sheet("Members")
        for r in dataframe_to_rows(df_members, index=False, header=True):
            ws_members.append(r)
        for cell in ws_members[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")

    # Sheet 3: Player Profiles
    if data["player_profiles"]:
        df_profiles = (
            pd.DataFrame(data["player_profiles"]) if data["player_profiles"] else pd.DataFrame()
        )
        profile_cols = {
            "id": "Player ID",
            "duprId": "DUPR ID",
            "fullName": "Full Name",
            "gender": "Gender",
            "age": "Age",
            "shortAddress": "Location",
            "singles": "Singles Rating",
            "doubles": "Doubles Rating",
            "singlesVerified": "Singles Verified",
            "doublesVerified": "Doubles Verified",
            "singlesReliability": "Singles Reliability (%)",
            "doublesReliability": "Doubles Reliability (%)",
            "singles_rating": "Singles Rating (Detailed)",
            "doubles_rating": "Doubles Rating (Detailed)",
        }
        available_cols = [c for c in profile_cols.keys() if c in df_profiles.columns]
        df_profiles = df_profiles[available_cols].rename(columns=profile_cols)

        ws_profiles = wb.create_sheet("Player Profiles")
        for r in dataframe_to_rows(df_profiles, index=False, header=True):
            ws_profiles.append(r)
        for cell in ws_profiles[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")

    # Sheet 4: Match History
    if data["match_history"]:
        match_data: List[Dict[str, Any]] = []
        for player_id, matches in data["match_history"].items():
            for match in matches:
                player_team = None
                opponent_team = None
                for team in match.get("teams", []):
                    if str(team.get("player1", {}).get("id")) == player_id or str(
                        team.get("player2", {}).get("id")
                    ) == player_id:
                        player_team = team
                    else:
                        opponent_team = team

                match_info: Dict[str, Any] = {
                    "player_id": player_id,
                    "match_id": match.get("id"),
                    "event_name": match.get("eventName"),
                    "event_date": match.get("eventDate"),
                    "event_format": match.get("eventFormat"),
                    "league": match.get("league"),
                    "tournament": match.get("tournament"),
                    "score_format": match.get("scoreFormat", {}).get("format")
                    if isinstance(match.get("scoreFormat"), dict)
                    else match.get("scoreFormat"),
                    "confirmed": match.get("confirmed"),
                }

                if player_team:
                    match_info.update(
                        {
                            "player_team_id": player_team.get("id"),
                            "player1_id": player_team.get("player1", {}).get("id"),
                            "player1_name": player_team.get("player1", {}).get("fullName"),
                            "player1_dupr": player_team.get("player1", {}).get("duprId"),
                            "player2_id": player_team.get("player2", {}).get("id"),
                            "player2_name": player_team.get("player2", {}).get("fullName"),
                            "player2_dupr": player_team.get("player2", {}).get("duprId"),
                            "player_team_winner": player_team.get("winner"),
                            "game1_score": player_team.get("game1"),
                            "game2_score": player_team.get("game2"),
                            "game3_score": player_team.get("game3"),
                            "player_team_rating_before": player_team.get(
                                "preMatchRatingAndImpact", {}
                            ).get(
                                "preMatchDoubleRatingPlayer1"
                            )
                            if match.get("eventFormat") == "DOUBLES"
                            else player_team.get("preMatchRatingAndImpact", {}).get(
                                "preMatchSingleRatingPlayer1"
                            ),
                            "player_team_rating_after": player_team.get("player1", {})
                            .get("postMatchRating", {})
                            .get("doubles")
                            if match.get("eventFormat") == "DOUBLES"
                            else player_team.get("player1", {})
                            .get("postMatchRating", {})
                            .get("singles"),
                        }
                    )

                if opponent_team:
                    match_info.update(
                        {
                            "opponent_team_id": opponent_team.get("id"),
                            "opponent1_id": opponent_team.get("player1", {}).get("id"),
                            "opponent1_name": opponent_team.get("player1", {}).get("fullName"),
                            "opponent1_dupr": opponent_team.get("player1", {}).get("duprId"),
                            "opponent2_id": opponent_team.get("player2", {}).get("id"),
                            "opponent2_name": opponent_team.get("player2", {}).get("fullName"),
                            "opponent2_dupr": opponent_team.get("player2", {}).get("duprId"),
                            "opponent_team_winner": opponent_team.get("winner"),
                            "opponent_game1_score": opponent_team.get("game1"),
                            "opponent_game2_score": opponent_team.get("game2"),
                            "opponent_game3_score": opponent_team.get("game3"),
                            "opponent_team_rating_before": opponent_team.get(
                                "preMatchRatingAndImpact", {}
                            ).get(
                                "preMatchDoubleRatingPlayer1"
                            )
                            if match.get("eventFormat") == "DOUBLES"
                            else opponent_team.get("preMatchRatingAndImpact", {}).get(
                                "preMatchSingleRatingPlayer1"
                            ),
                            "opponent_team_rating_after": opponent_team.get("player1", {})
                            .get("postMatchRating", {})
                            .get("doubles")
                            if match.get("eventFormat") == "DOUBLES"
                            else opponent_team.get("player1", {})
                            .get("postMatchRating", {})
                            .get("singles"),
                        }
                    )

                if player_team and opponent_team:
                    match_info["result"] = "Win" if player_team.get("winner") else "Loss"
                    match_info["score_summary"] = f"{player_team.get('game1')}-{opponent_team.get('game1')}"
                    if player_team.get("game2") != -1:
                        match_info["score_summary"] += (
                            f", {player_team.get('game2')}-{opponent_team.get('game2')}"
                        )

                match_data.append(match_info)

        df_matches = pd.DataFrame(match_data)
        match_cols = {
            "player_id": "Player ID",
            "match_id": "Match ID",
            "event_name": "Event Name",
            "event_date": "Match Date",
            "event_format": "Match Format",
            "league": "League",
            "tournament": "Tournament",
            "score_format": "Score Format",
            "result": "Result",
            "score_summary": "Score Summary",
            "confirmed": "Confirmed",
            "player1_name": "Player 1 Name",
            "player1_dupr": "Player 1 DUPR ID",
            "player2_name": "Player 2 Name",
            "player2_dupr": "Player 2 DUPR ID",
            "player_team_winner": "Player Team Won",
            "game1_score": "Player Game 1 Score",
            "game2_score": "Player Game 2 Score",
            "game3_score": "Player Game 3 Score",
            "player_team_rating_before": "Player Rating Before",
            "player_team_rating_after": "Player Rating After",
            "opponent1_name": "Opponent 1 Name",
            "opponent1_dupr": "Opponent 1 DUPR ID",
            "opponent2_name": "Opponent 2 Name",
            "opponent2_dupr": "Opponent 2 DUPR ID",
            "opponent_team_winner": "Opponent Team Won",
            "opponent_game1_score": "Opponent Game 1 Score",
            "opponent_game2_score": "Opponent Game 2 Score",
            "opponent_game3_score": "Opponent Game 3 Score",
            "opponent_team_rating_before": "Opponent Rating Before",
            "opponent_team_rating_after": "Opponent Rating After",
        }
        available_cols = [c for c in match_cols.keys() if c in df_matches.columns]
        df_matches = df_matches[available_cols].rename(columns=match_cols)

        ws_matches = wb.create_sheet("Match History")
        for r in dataframe_to_rows(df_matches, index=False, header=True):
            ws_matches.append(r)

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws_matches[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws_matches.freeze_panes = "A2"

    # Auto-fit columns (best effort)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{filename_prefix}_{ts}.xlsx"

    # Optional: embed charts into a new sheet if analysis_df exists globally at runtime
    try:
        # Rebuild a DataFrame resembling match history for charting
        all_rows: list[dict[str, Any]] = []
        for _, matches in data.get("match_history", {}).items():
            for m in matches:
                row = {
                    "Player Rating Before": m.get("player_team_rating_before"),
                    "Player Rating After": m.get("player_team_rating_after"),
                    "Opponent Rating Before": m.get("opponent_team_rating_before"),
                    "Result": m.get("result"),
                    "Match Date": m.get("event_date"),
                    "Match Format": m.get("event_format"),
                }
                all_rows.append(row)
        df_for_charts = pd.DataFrame(all_rows)
        if not df_for_charts.empty:
            img_sheet = wb.create_sheet("Charts")

            def _add_chart(img_data_uri: str, cell: str, title: str) -> None:
                try:
                    # Convert data URI to PIL and attach
                    base64_part = img_data_uri.split(",", 1)[1]
                    raw = base64.b64decode(base64_part)
                    bio = io.BytesIO(raw)
                    pil = PILImage.open(bio)
                    tmp = io.BytesIO()
                    pil.save(tmp, format="PNG")
                    tmp.seek(0)
                    xl_img = XLImage(tmp)
                    img_sheet.add_image(xl_img, cell)
                    img_sheet[cell].value = title
                except Exception:
                    pass

            # Generate and embed four charts
            _add_chart(render_chart_image("Rating distribution", df_for_charts, "All"), "A1", "Rating distribution")
            _add_chart(render_chart_image("Win/Loss breakdown", df_for_charts, "All"), "I1", "Win/Loss breakdown")
            _add_chart(render_chart_image("Diff vs Delta scatter", df_for_charts, "All"), "A25", "Diff vs Delta")
            _add_chart(render_chart_image("Matches over time", df_for_charts, "All"), "I25", "Matches over time")
            # Gender pie if gender info exists (from members/profiles sheets)
            if "Gender" in df_for_charts.columns or ("members" in data and data["members"]):
                # Try to merge gender from members if needed
                if "Gender" not in df_for_charts.columns and data.get("members"):
                    try:
                        df_members = pd.DataFrame(data["members"]).rename(columns={"gender": "Gender"})
                        if not df_members.empty and "Gender" in df_members.columns:
                            # Use members-only gender for the pie
                            _add_chart(render_chart_image("Gender pie", df_members[["Gender"]], "All"), "A49", "Gender pie")
                    except Exception:
                        pass
                else:
                    _add_chart(render_chart_image("Gender pie", df_for_charts, "All"), "A49", "Gender pie")
    except Exception:
        # If anything fails, still save workbook
        pass

    wb.save(filename)
    return filename


def main(page: ft.Page) -> None:
    page.title = "DUPR Club Crawler"
    page.window_width = 980
    page.window_height = 820
    page.padding = 16
    page.scroll = ft.ScrollMode.ALWAYS

    if _IMPORT_ERROR:
        page.add(
            ft.Text(
                "Không thể import dupr_client (duprly). Vui lòng cài đặt requirements trước.",
                color="red",
                weight=ft.FontWeight.BOLD,
            ),
            ft.Text(str(_IMPORT_ERROR), color="red"),
        )
        return

    email_input = ft.TextField(label="DUPR Email", value="", width=340)
    password_input = ft.TextField(label="DUPR Password", password=True, can_reveal_password=True, width=340)
    club_id_input = ft.TextField(label="Club ID", value="5986040853", width=200)
    max_members_input = ft.TextField(label="Max Members", value="20", width=150)
    max_matches_input = ft.TextField(label="Players to fetch history", value="10", width=200)
    matches_per_player_input = ft.TextField(label="Matches per player", value="20", width=200)
    filename_prefix_input = ft.TextField(label="Filename prefix", value="dupr_club", width=200)

    start_button = ft.ElevatedButton(text="Start Crawl & Export", disabled=False)
    stop_button = ft.ElevatedButton(text="Stop", disabled=True)
    login_button = ft.ElevatedButton(text="Login")
    guest_login_button = ft.OutlinedButton(text="Login as Guest")
    colab_button = ft.TextButton(text="Open Colab Script")
    dupr_button = ft.TextButton(text="Open DUPR.com")
    help_button = ft.OutlinedButton(text="Help")
    progress = ft.ProgressBar(width=400, visible=False)
    status_text = ft.Text(value="", color="blue", selectable=True)
    log_list = ft.ListView(expand=0, height=100, spacing=2, auto_scroll=True)
    logs_container = ft.Container(content=log_list, width=800, padding=8, border=ft.border.all(1, "#EEEEEE"), border_radius=8)

    # ===== Rating Analysis & Simulation state =====
    analysis_df: Optional[pd.DataFrame] = None
    analysis_profiles_df: Optional[pd.DataFrame] = None
    fitted_K: Optional[float] = None
    fitted_scale: Optional[float] = None
    mae_text = ft.Text(visible=False, selectable=True)
    formula_img1 = ft.Image(visible=False)
    formula_img2 = ft.Image(visible=False)
    formula_img3 = ft.Image(visible=False)
    formula_img3 = ft.Image(visible=False)
    formula_text_container = ft.Container(
        visible=False,
        bgcolor="#263238",
        padding=12,
        border_radius=8,
        content=ft.Column([
            ft.Text(
                "rating_after = rating_before + K * (result - expected)",
                color="#FFFFFF",
                size=18,
                weight=ft.FontWeight.BOLD,
            ),
            ft.Text(
                "expected = 1 / (1 + 10^(-(diff/scale)))",
                color="#FFFFFF",
                size=18,
                weight=ft.FontWeight.BOLD,
            ),
            ft.Text(
                "diff = rating_before - opponent_rating_before",
                color="#FFFFFF",
                size=18,
                weight=ft.FontWeight.BOLD,
            ),
        ], spacing=4),
    )

    # Copyable, pretty math text blocks (hidden until Excel analysis is done)
    formula_plain_text = ft.Text(value="", color="#FFFFFF", size=16, selectable=True, style=ft.TextStyle(font_family="monospace"))
    formula_latex_text = ft.Text(value="", color="#FFFFFF", size=16, selectable=True, style=ft.TextStyle(font_family="monospace"))

    def copy_to_clipboard(text: str) -> None:
        try:
            page.set_clipboard(text)
            set_status("Copied to clipboard", "green")
        except Exception:
            pass

    formula_plain_container = ft.Container(
        visible=False,
        bgcolor="#263238",
        padding=12,
        border_radius=8,
        content=ft.Column([
            ft.Row([
                ft.Text("Plain", color="#BBDEFB", weight=ft.FontWeight.BOLD),
                ft.TextButton("Copy", on_click=lambda _: copy_to_clipboard(formula_plain_text.value)),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            formula_plain_text,
        ], spacing=6),
    )

    formula_latex_container = ft.Container(
        visible=False,
        bgcolor="#263238",
        padding=12,
        border_radius=8,
        content=ft.Column([
            ft.Row([
                ft.Text("LaTeX", color="#BBDEFB", weight=ft.FontWeight.BOLD),
                ft.TextButton("Copy", on_click=lambda _: copy_to_clipboard(formula_latex_text.value)),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            formula_latex_text,
        ], spacing=6),
    )
    sample_dropdown = ft.Dropdown(label="Chọn ví dụ từ Excel (điền form tự động)", width=600)

    # ===== Visualization controls =====
    visual_help_panel = ft.Ref[ft.Container]()
    chart_type_dd = ft.Dropdown(
        label="Biểu đồ",
        width=150,
        options=[
            ft.dropdown.Option("Rating distribution"),
            ft.dropdown.Option("Win/Loss breakdown"),
            ft.dropdown.Option("Diff vs Delta scatter"),
            ft.dropdown.Option("Matches over time"),
            ft.dropdown.Option("Gender pie"),
        ],
        value="Rating distribution",
        visible=False,
    )
    format_filter_dd = ft.Dropdown(
        label="Format",
        width=180,
        options=[ft.dropdown.Option("All"), ft.dropdown.Option("Singles"), ft.dropdown.Option("Doubles")],
        value="All",
        visible=False,
    )
    update_chart_btn = ft.ElevatedButton(text="Update Chart", visible=False)
    chart_img = ft.Image(visible=False, width=560, height=340)

    def on_toggle_visual_help(_: ft.ControlEvent) -> None:
        if visual_help_panel.current is None:
            return
        visual_help_panel.current.visible = not visual_help_panel.current.visible
        page.update()

    file_picker = ft.FilePicker()
    page.overlay.append(file_picker)

    def compute_expected(prob_diff: float, scale: float) -> float:
        return 1.0 / (1.0 + 10 ** (-(prob_diff / scale)))

    def fit_parameters(df: pd.DataFrame) -> tuple[Optional[float], Optional[float], Optional[float], int]:
        try:
            needed = [
                "Player Rating Before",
                "Player Rating After",
                "Opponent Rating Before",
                "Result",
            ]
            for col in needed:
                if col not in df.columns:
                    return None, None, None, 0

            df_use = df.copy()
            df_use["y"] = df_use["Result"].map({"Win": 1, "Loss": 0})
            df_use = df_use.dropna(
                subset=[
                    "Player Rating Before",
                    "Player Rating After",
                    "Opponent Rating Before",
                    "y",
                ]
            )
            if df_use.empty:
                return None, None, None, 0

            df_use["delta"] = df_use["Player Rating After"] - df_use["Player Rating Before"]
            df_use["diff"] = df_use["Player Rating Before"] - df_use["Opponent Rating Before"]

            best_scale = None
            best_mae = float("inf")
            best_K = None

            for scale in range(200, 901, 50):
                expected = 1.0 / (1.0 + np.power(10.0, -(df_use["diff"].values / float(scale))))
                numerator = df_use["delta"].values
                denom = (df_use["y"].values - expected)
                mask = np.abs(denom) > 1e-6
                if not np.any(mask):
                    continue
                K_i = numerator[mask] / denom[mask]
                K = float(np.median(K_i))
                pred_delta = K * (df_use["y"].values - expected)
                mae = float(np.mean(np.abs(pred_delta - df_use["delta"].values)))
                if mae < best_mae:
                    best_mae = mae
                    best_scale = float(scale)
                    best_K = K

            if best_scale is None or best_K is None:
                return None, None, None, 0

            return best_K, best_scale, best_mae, int(len(df_use))
        except Exception:
            return None, None, None, 0

    def build_samples_dropdown(df: pd.DataFrame) -> None:
        options: List[ft.dropdown.Option] = []
        if df is None or df.empty:
            sample_dropdown.options = options
            return
        cols = df.columns
        # Build from first 50 valid rows
        count = 0
        for idx, row in df.iterrows():
            try:
                p_before = float(row.get("Player Rating Before"))
                o_before = float(row.get("Opponent Rating Before"))
            except Exception:
                continue
            res = str(row.get("Result", "")).title()
            p1 = row.get("Player 1 Name", "") if "Player 1 Name" in cols else ""
            opp1 = row.get("Opponent 1 Name", "") if "Opponent 1 Name" in cols else ""
            date = row.get("Match Date", "") if "Match Date" in cols else ""
            label_parts = []
            if p1 or opp1:
                label_parts.append(f"{p1} vs {opp1}")
            if date:
                label_parts.append(str(date))
            label_parts.append(f"{p_before:.2f} vs {o_before:.2f}")
            if res in ("Win", "Loss"):
                label_parts.append(f"Result: {res}")
            label = " | ".join([str(x) for x in label_parts if str(x)])
            options.append(ft.dropdown.Option(text=label, key=str(idx)))
            count += 1
            if count >= 50:
                break
        sample_dropdown.options = options

    def _latex_tools_available() -> bool:
        # Matplotlib usetex needs latex and dvipng in PATH
        return shutil.which("latex") is not None and shutil.which("dvipng") is not None

    def render_formula_image(latex_str: str) -> str:
        use_usetex = _latex_tools_available()
        if use_usetex:
            # High-quality LaTeX rendering using system LaTeX (if installed)
            with matplotlib.rc_context({
                "text.usetex": True,
                "font.family": "serif",
                "text.latex.preamble": r"\usepackage{amsmath}\usepackage{amssymb}",
            }):
                fig, ax = plt.subplots(figsize=(6.8, 1.8))
                ax.axis("off")
                fig.patch.set_alpha(0.0)
                ax.text(0.02, 0.5, latex_str, fontsize=24, ha="left", va="center")
                buf = io.BytesIO()
                fig.savefig(
                    buf,
                    format="png",
                    bbox_inches="tight",
                    pad_inches=0.12,
                    dpi=280,
                    transparent=True,
                )
                plt.close(fig)
        else:
            # Fallback to mathtext (no external LaTeX required)
            fig, ax = plt.subplots(figsize=(6.8, 1.8))
            ax.axis("off")
            fig.patch.set_alpha(0.0)
            ax.text(0.02, 0.5, latex_str, fontsize=20, ha="left", va="center")
            buf = io.BytesIO()
            fig.savefig(buf, format="png", bbox_inches="tight", pad_inches=0.12, dpi=240, transparent=True)
            plt.close(fig)
        data = base64.b64encode(buf.getvalue()).decode("ascii")
        return f"data:image/png;base64,{data}"

    def render_chart_image(kind: str, df: pd.DataFrame, fmt: str = "All") -> str:
        try:
            plot_df = df.copy()
            if fmt in ("Singles", "Doubles") and "Match Format" in plot_df.columns:
                plot_df = plot_df[plot_df["Match Format"].astype(str).str.title() == fmt]

            fig, ax = plt.subplots(figsize=(5.6, 3.4))
            fig.patch.set_alpha(0.0)
            ax.set_facecolor("#263238")

            if kind == "Rating distribution":
                cols = [c for c in ["Player Rating Before", "Opponent Rating Before"] if c in plot_df.columns]
                vals = pd.concat([plot_df[c] for c in cols], ignore_index=True) if cols else pd.Series([], dtype=float)
                vals = pd.to_numeric(vals, errors="coerce").dropna()
                n, bins, patches = ax.hist(vals, bins=25, color="#42A5F5", edgecolor="#1E88E5")
                ax.set_xlabel("Rating", color="#FFFFFF")
                ax.set_ylabel("Count", color="#FFFFFF")
                ax.set_title("Rating distribution", color="#FFFFFF")
                # show bin counts on top bars
                for i in range(len(n)):
                    if n[i] > 0:
                        x = (bins[i] + bins[i+1]) / 2
                        ax.annotate(f"{int(n[i])}", xy=(x, n[i]), xytext=(0, 3), textcoords="offset points",
                                    ha="center", va="bottom", color="#FFFFFF", fontsize=8)
                ax.tick_params(colors="#FFFFFF")
                for spine in ax.spines.values():
                    spine.set_color("#FFFFFF")

            elif kind == "Win/Loss breakdown":
                if "Result" in plot_df.columns:
                    counts = plot_df["Result"].value_counts()
                    bars = ax.bar(counts.index, counts.values, color=["#66BB6A" if i=="Win" else "#EF5350" for i in counts.index])
                    # annotate counts on top of bars
                    for b in bars:
                        height = b.get_height()
                        ax.annotate(f"{int(height)}", xy=(b.get_x()+b.get_width()/2, height), xytext=(0, 3),
                                    textcoords="offset points", ha="center", va="bottom", color="#FFFFFF", fontsize=10, fontweight="bold")
                    ax.set_title("Win/Loss breakdown", color="#FFFFFF")
                    ax.set_xlabel("Result", color="#FFFFFF")
                    ax.set_ylabel("Count", color="#FFFFFF")
                    ax.tick_params(colors="#FFFFFF")
                    for spine in ax.spines.values():
                        spine.set_color("#FFFFFF")
                else:
                    ax.text(0.5, 0.5, "No 'Result' column", ha="center", va="center")

            elif kind == "Diff vs Delta scatter":
                required = ["Player Rating Before", "Opponent Rating Before", "Player Rating After"]
                if all(c in plot_df.columns for c in required):
                    plot_df = plot_df.copy()
                    plot_df["diff"] = plot_df["Player Rating Before"] - plot_df["Opponent Rating Before"]
                    plot_df["delta"] = plot_df["Player Rating After"] - plot_df["Player Rating Before"]
                    sc = ax.scatter(plot_df["diff"], plot_df["delta"], s=14, alpha=0.6, color="#8E24AA")
                    ax.set_xlabel("diff (player - opponent)", color="#FFFFFF")
                    ax.set_ylabel("delta (rating change)", color="#FFFFFF")
                    ax.set_title("Diff vs Delta", color="#FFFFFF")
                    ax.tick_params(colors="#FFFFFF")
                    for spine in ax.spines.values():
                        spine.set_color("#FFFFFF")
                else:
                    ax.text(0.5, 0.5, "Columns missing", ha="center", va="center")

            elif kind == "Matches over time":
                # Count matches per date
                if "Match Date" in plot_df.columns:
                    try:
                        dates = pd.to_datetime(plot_df["Match Date"], errors="coerce")
                        counts = dates.dt.date.value_counts().sort_index()
                        ax.plot(list(counts.index), list(counts.values), marker="o", color="#3949AB")
                        ax.set_xlabel("Date", color="#FFFFFF")
                        ax.set_ylabel("Matches", color="#FFFFFF")
                        ax.set_title("Matches over time", color="#FFFFFF")
                        ax.tick_params(colors="#FFFFFF")
                        for spine in ax.spines.values():
                            spine.set_color("#FFFFFF")
                        fig.autofmt_xdate()
                    except Exception:
                        ax.text(0.5, 0.5, "Invalid dates", ha="center", va="center")
                else:
                    ax.text(0.5, 0.5, "No 'Match Date' column", ha="center", va="center")

            elif kind == "Gender pie":
                # Prefer Player Profiles sheet gender if loaded
                gender_series = None
                try:
                    if analysis_profiles_df is not None and not analysis_profiles_df.empty and "Gender" in analysis_profiles_df.columns:
                        gender_series = analysis_profiles_df["Gender"]
                except Exception:
                    pass
                # Fallbacks
                if gender_series is None and "Gender" in plot_df.columns:
                    gender_series = plot_df["Gender"]
                if gender_series is None and "Gender" in df.columns:
                    gender_series = df["Gender"]
                if gender_series is not None:
                    g = gender_series.astype(str).str.strip().str.upper()
                    counts = g.replace({"M": "MALE", "F": "FEMALE", "Male": "MALE", "Female": "FEMALE"})
                    counts = counts[counts.isin(["MALE", "FEMALE"])].value_counts()
                    if not counts.empty:
                        def _autopct(pct):
                            total = counts.sum()
                            val = int(round(pct*total/100.0))
                            return f"{pct:.1f}%\n({val})"
                        wedges, texts, autotexts = ax.pie(
                            counts.values,
                            labels=counts.index,
                            autopct=_autopct,
                            textprops={"color": "#FFFFFF"},
                            colors=["#42A5F5", "#EF5350"],
                        )
                        ax.set_title("Gender pie", color="#FFFFFF")
                    else:
                        ax.text(0.5, 0.5, "No Male/Female data", ha="center", va="center")
                else:
                    ax.text(0.5, 0.5, "No Gender column", ha="center", va="center")

            ax.grid(True, alpha=0.25)
            buf = io.BytesIO()
            fig.savefig(buf, format="png", bbox_inches="tight", dpi=200)
            plt.close(fig)
            data = base64.b64encode(buf.getvalue()).decode("ascii")
            return f"data:image/png;base64,{data}"
        except Exception as _:
            # Return small blank image to avoid UI crash
            fig, ax = plt.subplots(figsize=(1, 1))
            ax.axis("off")
            buf = io.BytesIO()
            fig.savefig(buf, format="png")
            plt.close(fig)
            data = base64.b64encode(buf.getvalue()).decode("ascii")
            return f"data:image/png;base64,{data}"

    def _set_image_from_data_uri(img: ft.Image, data_uri: str) -> None:
        try:
            if isinstance(data_uri, str) and data_uri.startswith("data:image") and "," in data_uri:
                img.src = None
                img.src_base64 = data_uri.split(",", 1)[1]
            else:
                img.src_base64 = None
                img.src = data_uri
        except Exception:
            # As a last resort, hide image to avoid runtime error
            img.visible = False

    
    
    

    def on_pick_result(e: ft.FilePickerResultEvent) -> None:
        nonlocal analysis_df, analysis_profiles_df, fitted_K, fitted_scale
        if not e.files:
            set_status("Không chọn file.")
            return
        path = e.files[0].path or ""
        if not path.lower().endswith(".xlsx"):
            set_status("Vui lòng chọn file Excel .xlsx", "red")
            return
        try:
            append_log(f"Đang đọc file: {path}")
            xls = None
            try:
                dfm = pd.read_excel(path, sheet_name="Match History")
            except Exception:
                xls = pd.ExcelFile(path)
                sheet = "Match History" if "Match History" in xls.sheet_names else xls.sheet_names[0]
                dfm = pd.read_excel(xls, sheet_name=sheet)
            col_map = {}
            if "Player Rating Before" not in dfm.columns and "player_team_rating_before" in dfm.columns:
                col_map["player_team_rating_before"] = "Player Rating Before"
            if "Player Rating After" not in dfm.columns and "player_team_rating_after" in dfm.columns:
                col_map["player_team_rating_after"] = "Player Rating After"
            if "Opponent Rating Before" not in dfm.columns and "opponent_team_rating_before" in dfm.columns:
                col_map["opponent_team_rating_before"] = "Opponent Rating Before"
            if "Result" not in dfm.columns and "result" in dfm.columns:
                col_map["result"] = "Result"
            if col_map:
                dfm = dfm.rename(columns=col_map)

            # Normalize dtypes and values
            for c in ["Player Rating Before", "Player Rating After", "Opponent Rating Before"]:
                if c in dfm.columns:
                    dfm[c] = pd.to_numeric(dfm[c], errors="coerce")
            if "Result" in dfm.columns:
                dfm["Result"] = dfm["Result"].astype(str).str.strip().str.title()
                dfm["Result"] = dfm["Result"].replace({
                    "True": "Win",
                    "False": "Loss",
                    "1": "Win",
                    "0": "Loss",
                })

            # Also try to load Player Profiles sheet for gender analysis
            analysis_profiles_df = None
            try:
                if xls is None:
                    # Open once if not already
                    xls = pd.ExcelFile(path)
                if "Player Profiles" in xls.sheet_names:
                    df_profiles = pd.read_excel(xls, sheet_name="Player Profiles")
                    # Normalize column names minimally
                    if "Gender" not in df_profiles.columns and "gender" in df_profiles.columns:
                        df_profiles = df_profiles.rename(columns={"gender": "Gender"})
                    analysis_profiles_df = df_profiles
            except Exception:
                analysis_profiles_df = None

            analysis_df = dfm
            K, scale, mae, n = fit_parameters(analysis_df)
            fitted_K, fitted_scale = K, scale
            if K is None or scale is None:
                mae_text.value = "Không thể ước lượng tham số từ dữ liệu."
                mae_text.visible = True
                formula_img1.visible = False
                formula_img2.visible = False
                formula_text_container.visible = False
                try:
                    formula_plain_container.visible = False
                    formula_latex_container.visible = False
                except Exception:
                    pass
            else:
                # Render math images (manual LaTeX to ensure mathtext compatibility)
                img1 = render_formula_image(r"$rating_{after} = rating_{before} + K\,(y - E)$")
                img2 = render_formula_image(r"$E = \frac{1}{1 + 10^{-(\Delta/\mathrm{scale})}},\ \ \Delta = rating_{before} - rating^{opp}_{before}$")
                formula_img1.src = img1
                formula_img2.src = img2
                formula_img1.visible = True
                formula_img2.visible = True
                formula_text_container.visible = True
                # Update copyable formula blocks
                try:
                    formula_plain_text.value = (
                        "rating_after = rating_before + K * (result - expected)\n"
                        "expected = 1 / (1 + 10^(-(diff/scale)))\n"
                        "diff = rating_before - opponent_rating_before"
                    )
                    formula_latex_text.value = (
                        r"rating\_{after} = rating\_{before} + K\,(y - E)\n"
                        r"E = \frac{1}{1 + 10^{-\,(\Delta/scale)}}\n"
                        r"\Delta = rating\_{before} - rating^{opp}\_{before}"
                    )
                    formula_plain_container.visible = True
                    formula_latex_container.visible = True
                except Exception:
                    pass
                mae_text.value = (
                    f"- K ≈ {K:.4f}\n"
                    f"- scale ≈ {scale:.1f}\n"
                    f"- MAE ≈ {mae:.4f} (n={n})\n"
                )
                mae_text.visible = True
            build_samples_dropdown(analysis_df)
            # Show chart controls when data is ready
            chart_type_dd.visible = True
            format_filter_dd.visible = True
            update_chart_btn.visible = True
            # Auto-render a default chart to ensure Image has a source
            try:
                default_kind = chart_type_dd.value or "Rating distribution"
                default_fmt = format_filter_dd.value or "All"
                _uri = render_chart_image(default_kind, analysis_df, default_fmt)
                _set_image_from_data_uri(chart_img, _uri)
                chart_img.visible = True
            except Exception:
                chart_img.visible = False
            set_status("Đọc & ước lượng xong", "green")
            page.update()
        except Exception as exc:
            set_status(f"Lỗi đọc/analysis file: {exc}", "red")

    file_picker.on_result = on_pick_result

    # Do not show formulas until Excel loaded

    match_format = ft.Dropdown(
        label="Format",
        width=150,
        options=[ft.dropdown.Option("Singles"), ft.dropdown.Option("Doubles")],
        value="Singles",
    )
    sim_player_before = ft.TextField(label="Player rating before", width=180)
    sim_teammate_before = ft.TextField(label="Teammate rating before (doubles)", width=220, visible=False)
    sim_opp_before = ft.TextField(label="Opponent rating before", width=180)
    sim_opp2_before = ft.TextField(label="Opponent 2 rating before (doubles)", width=240, visible=False)
    sim_scores = ft.TextField(
        label="Game scores (e.g., 11-9,7-11,11-6)",
        width=360,
        hint_text="Optional: determines Win/Loss from games if provided",
    )
    sim_result = ft.Dropdown(
        label="Result",
        width=150,
        options=[ft.dropdown.Option("Win"), ft.dropdown.Option("Loss")],
        value="Win",
    )
    sim_example_btn = ft.OutlinedButton(text="Run Example")
    sim_compute_btn = ft.ElevatedButton(text="Compute Simulation")

    def on_update_chart(_: ft.ControlEvent) -> None:
        if analysis_df is None or analysis_df.empty:
            set_status("Chưa có dữ liệu để vẽ biểu đồ", "red")
            return
        kind = chart_type_dd.value or "Rating distribution"
        fmt = format_filter_dd.value or "All"
        try:
            # For Gender pie, prefer Player Profiles sheet if loaded
            _df = analysis_df
            if kind == "Gender pie" and isinstance(analysis_df, pd.DataFrame):
                pass
            _uri = render_chart_image(kind, _df, fmt)
            _set_image_from_data_uri(chart_img, _uri)
            chart_img.visible = True
        except Exception as exc:
            set_status(f"Lỗi vẽ biểu đồ: {exc}", "red")
            chart_img.visible = False
        page.update()

    update_chart_btn.on_click = on_update_chart

    # Result UI components (clearer, more emphasized)
    result_main_value = ft.Text(size=30, weight=ft.FontWeight.W_700, color="#000000", selectable=True)
    result_main_label = ft.Text("New rating", size=12, color="#000000", weight=ft.FontWeight.BOLD, selectable=True)
    pill_delta_text = ft.Text(weight=ft.FontWeight.BOLD, color="#000000", selectable=True)
    pill_delta = ft.Container(
        content=pill_delta_text,
        padding=8,
        border_radius=999,
        bgcolor="#E8F5E9",
        border=ft.border.all(1, "#2E7D32"),
    )
    pill_expected_text = ft.Text(weight=ft.FontWeight.BOLD, color="#000000", selectable=True)
    pill_expected = ft.Container(
        content=pill_expected_text,
        padding=8,
        border_radius=999,
        bgcolor="#E3F2FD",
        border=ft.border.all(1, "#1565C0"),
    )
    detail_list = ft.Column(spacing=2)

    result_container = ft.Container(
        visible=False,
        padding=14,
        border=ft.border.all(1, "#E0E0E0"),
        border_radius=10,
        bgcolor="#FFFFFF",
        content=ft.Column([
            ft.Text("Simulation Result", size=16, weight=ft.FontWeight.BOLD, color="#000000", selectable=True),
            ft.Row([
                ft.Container(
                    content=ft.Column([result_main_label, result_main_value], spacing=2),
                    padding=12,
                    bgcolor="#FFFFFF",
                    border=ft.border.all(1, "#E0E0E0"),
                    border_radius=10,
                ),
                ft.Row([pill_delta, pill_expected], spacing=10),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            detail_list,
        ], spacing=10),
    )

    def get_params() -> tuple[float, float, bool]:
        # Return (K, scale, is_fallback)
        if fitted_K is not None and fitted_scale is not None:
            return float(fitted_K), float(fitted_scale), False
        # Fallback demo parameters
        return 0.10, 400.0, True

    def run_compute(a_team: float, b_team: float, y_val: float) -> None:
        K, scale, is_fallback = get_params()
        expected = compute_expected(a_team - b_team, scale)
        delta = K * (y_val - expected)
        after = a_team + delta

        # Main value and pills
        result_main_value.value = f"{after:.4f}"
        pill_delta_text.value = f"Δ {delta:+.4f}"
        pill_expected_text.value = f"E {expected:.4f}"

        # Delta coloring
        if delta >= 0:
            pill_delta.bgcolor = "#E8F5E9"
            pill_delta.border = ft.border.all(1, "#2E7D32")
        else:
            pill_delta.bgcolor = "#FFF3E0"
            pill_delta.border = ft.border.all(1, "#EF6C00")

        # Details list
        detail_list.controls = [
            ft.Text(f"Before: {a_team:.4f}", weight=ft.FontWeight.BOLD, color="#000000", selectable=True),
            ft.Text(f"Opponent before: {b_team:.4f}", weight=ft.FontWeight.BOLD, color="#000000", selectable=True),
            ft.Text(f"Diff: {a_team - b_team:.4f}", weight=ft.FontWeight.BOLD, color="#000000", selectable=True),
            ft.Text(f"Result: {'Win' if y_val == 1.0 else 'Loss'}", weight=ft.FontWeight.BOLD, color="#000000", selectable=True),
            ft.Text(f"K: {K:.4f}   |   scale: {scale:.1f} {'(demo)' if is_fallback else ''}", weight=ft.FontWeight.BOLD, color="#000000", selectable=True),
        ]
        result_container.visible = True
        page.update()

    def on_sample_change(_: ft.ControlEvent) -> None:
        nonlocal analysis_df
        try:
            if analysis_df is None or analysis_df.empty:
                set_status("Chưa có dữ liệu. Hãy Load Excel trước.", "red")
                return
            sel = sample_dropdown.value
            if not sel:
                return
            row_idx = int(sel)
            if row_idx not in analysis_df.index:
                set_status("Ví dụ không hợp lệ.", "red")
                return
            row = analysis_df.loc[row_idx]
            a = float(row.get("Player Rating Before"))
            b = float(row.get("Opponent Rating Before"))
            res_str = str(row.get("Result", "")).title()
            sim_player_before.value = f"{a:.4f}"
            sim_opp_before.value = f"{b:.4f}"
            sim_result.value = "Win" if res_str == "Win" else "Loss"
            page.update()
        except Exception as exc:
            set_status(f"Không thể nạp ví dụ: {exc}", "red")

    def parse_scores(scores_str: str) -> Optional[float]:
        try:
            if not scores_str.strip():
                return None
            games = [g.strip() for g in scores_str.split(',') if g.strip()]
            if not games:
                return None
            wins_a = 0
            wins_b = 0
            for g in games:
                if '-' not in g:
                    continue
                a_s, b_s = g.split('-', 1)
                a_sc = int(a_s.strip())
                b_sc = int(b_s.strip())
                if a_sc > b_sc:
                    wins_a += 1
                elif b_sc > a_sc:
                    wins_b += 1
            if wins_a == wins_b:
                return None
            return 1.0 if wins_a > wins_b else 0.0
        except Exception:
            return None

    def on_format_change(_: ft.ControlEvent) -> None:
        is_doubles = (match_format.value == "Doubles")
        sim_teammate_before.visible = is_doubles
        sim_opp2_before.visible = is_doubles
        page.update()

    def on_simulate(_: ft.ControlEvent) -> None:
        try:
            fmt = match_format.value or "Singles"
            a_player = float(sim_player_before.value)
            b_player = float(sim_opp_before.value)
            if fmt == "Doubles":
                a_mate = float(sim_teammate_before.value)
                b_mate = float(sim_opp2_before.value)
                a_team = (a_player + a_mate) / 2.0
                b_team = (b_player + b_mate) / 2.0
            else:
                a_team = a_player
                b_team = b_player

            y_from_scores = parse_scores(sim_scores.value)
            y = y_from_scores if y_from_scores is not None else (1.0 if (sim_result.value or "Win") == "Win" else 0.0)
            run_compute(a_team, b_team, y)
        except Exception as exc:
            set_status(f"Input không hợp lệ: {exc}", "red")

    def on_example(_: ft.ControlEvent) -> None:
        # Example: Player 4.25 vs Opponent 4.75, player Wins
        sim_player_before.value = "4.25"
        sim_opp_before.value = "4.75"
        sim_result.value = "Win"
        page.update()
        run_compute(4.25, 4.75, 1.0)

    match_format.on_change = on_format_change
    sim_compute_btn.on_click = on_simulate
    sim_example_btn.on_click = on_example
    sample_dropdown.on_change = on_sample_change

    load_excel_btn = ft.ElevatedButton(
        text="Load Excel for Analysis",
        on_click=lambda _: file_picker.pick_files(allow_multiple=False),
    )

    def append_log(message: str) -> None:
        log_list.controls.append(ft.Text(message, selectable=True))
        page.update()

    def set_status(message: str, color: Optional[str] = None) -> None:
        status_text.value = message
        if color:
            status_text.color = color
        page.update()

    def set_running(is_running: bool) -> None:
        start_button.disabled = is_running
        progress.visible = is_running
        page.update()

    current_thread: Optional[threading.Thread] = None
    stop_event = threading.Event()
    client_instance: Optional[DuprClient] = None

    def perform_login(email: str, password: str) -> None:
        nonlocal client_instance
        try:
            if not email or not password:
                set_status("Thiếu email/password", "red")
                return
            if client_instance is None:
                client_instance = DuprClient()
            client_instance.email = email
            client_instance.password = password
            append_log("Đăng nhập DUPR…")
            client_instance.login_user(email, password)
            set_status("Đăng nhập thành công", "green")
        except Exception as exc:
            append_log(f"Lỗi đăng nhập: {exc}")
            set_status("Đăng nhập thất bại", "red")

    def do_crawl() -> None:
        try:
            set_running(True)
            stop_button.disabled = False
            page.update()
            append_log("Bắt đầu quá trình crawl dữ liệu…")

            email = email_input.value.strip()
            password = password_input.value
            club_id = club_id_input.value.strip()

            max_members = int(max_members_input.value or 0)
            max_matches = int(max_matches_input.value or 0)
            matches_per_player = int(matches_per_player_input.value or 0)
            filename_prefix = filename_prefix_input.value.strip() or "dupr_export"

            if not email or not password or not club_id:
                set_status("Thiếu thông tin email, password hoặc club ID", "red")
                return

            # Use existing logged-in client if available, otherwise create & login
            client = client_instance if client_instance is not None else DuprClient()
            client.email = email
            client.password = password
            append_log("Đăng nhập DUPR…")
            client.login_user(email, password)
            set_status("Đăng nhập thành công", "green")

            data = get_club_data(
                client,
                club_id,
                max_members=max_members,
                max_matches=max_matches,
                matches_per_player=matches_per_player,
                log=append_log,
                stop_event=stop_event,
            )

            append_log("Xuất dữ liệu ra Excel…")
            filename = export_to_excel(data, f"{filename_prefix}_{club_id}")
            full_path = os.path.abspath(filename)
            append_log(f"Xuất file thành công: {full_path}")
            set_status("Hoàn thành!", "green")

            # Offer to open file/folder
            try:
                if os.name == "nt":
                    os.startfile(full_path)  # type: ignore[attr-defined]
            except Exception:
                pass

        except Exception as exc:
            append_log(f"Lỗi: {exc}")
            set_status("Có lỗi xảy ra", "red")
        finally:
            set_running(False)
            stop_button.disabled = True
            page.update()

    def on_start_click(_: ft.ControlEvent) -> None:
        log_list.controls.clear()
        page.update()
        stop_event.clear()
        nonlocal current_thread
        current_thread = threading.Thread(target=do_crawl, daemon=True)
        current_thread.start()

    def on_stop_click(_: ft.ControlEvent) -> None:
        stop_event.set()
        append_log("Đang dừng… chờ tác vụ hiện tại kết thúc.")
        stop_button.disabled = True
        page.update()

    def on_login_click(_: ft.ControlEvent) -> None:
        perform_login(email_input.value.strip(), password_input.value)

    def on_guest_login_click(_: ft.ControlEvent) -> None:
        email_input.value = "duprsport@gmail.com"
        password_input.value = "crawldata123"
        page.update()
        perform_login(email_input.value.strip(), password_input.value)

    def on_colab_click(_: ft.ControlEvent) -> None:
        page.launch_url("https://colab.research.google.com/drive/1i3-P56YY819WM1foFppVXzwDqEjbZCQs?authuser=1#scrollTo=VqOJWMLXWGSR")

    def on_dupr_click(_: ft.ControlEvent) -> None:
        page.launch_url("https://www.dupr.com")

    # Help panel toggle
    help_panel = ft.Container(
        visible=False,
        padding=12,
        border=ft.border.all(1, "#E0E0E0"),
        border_radius=8,
        bgcolor="#FAFAFA",
        content=ft.Column([
            ft.Text("Giải thích tham số", weight=ft.FontWeight.BOLD, selectable=True, color="#000000"),
            ft.Text("- Club ID: Mã định danh câu lạc bộ trên DUPR.", selectable=True, color='#000000'),
            ft.Text("- Max Members: Số thành viên tối đa sẽ lấy thông tin từ club.", selectable=True, color='#000000'),
            ft.Text("- Players to fetch history: Số thành viên (trong danh sách) sẽ lấy lịch sử trận.", selectable=True, color='#000000'),
            ft.Text("- Matches per player: Số trận tối đa lấy cho mỗi thành viên.", selectable=True, color='#000000'),
            ft.Text("- Filename prefix: Tiền tố tên file Excel khi xuất.", selectable=True, color='#000000'),
            ft.Divider(),
            ft.Text("Ví dụ minh họa", weight=ft.FontWeight.BOLD, selectable=True, color='#000000'),
            ft.Text("- Club ID: https://dashboard.dupr.com/dashboard/browse/clubs/XXXXXXXXXX", selectable=True, color='#000000'),
            ft.Text("- Max Members: 20 (lấy thông tin 20 thành viên đầu tiên)", selectable=True, color='#000000'),
            ft.Text("- Players to fetch history: 10 (lấy lịch sử cho 10 người)", selectable=True, color='#000000'),
            ft.Text("- Matches per player: 20 (tối đa 20 trận mỗi người)", selectable=True, color='#000000'),
            ft.Text("- Filename prefix: dupr_club → dupr_club_<ClubID>_YYYYMMDD_HHMMSS.xlsx", selectable=True, color='#000000'),
        ], spacing=4),
    )

    def on_help_click(_: ft.ControlEvent) -> None:
        help_panel.visible = not help_panel.visible
        page.update()

    start_button.on_click = on_start_click
    stop_button.on_click = on_stop_click
    login_button.on_click = on_login_click
    guest_login_button.on_click = on_guest_login_click
    colab_button.on_click = on_colab_click
    dupr_button.on_click = on_dupr_click
    help_button.on_click = on_help_click

    form_row1 = ft.Row([email_input, password_input], spacing=16)
    form_row2 = ft.Row(
        [club_id_input, max_members_input, max_matches_input, matches_per_player_input, filename_prefix_input],
        wrap=True,
        spacing=16,
    )

    # Header with instructions at top-right
    instructions_chip = ft.Container(
        content=ft.Text("Hướng dẫn: 1) Login Guest & Start Crawl  2) Load Excel  3) Chọn Sample từ Excel  4) Compute Score", color="#FFFFFF", size=14, weight=ft.FontWeight.BOLD, selectable=True),
        padding=8,
        bgcolor="#1976D2",
        border_radius=8,
    )

    header_right = ft.Row([instructions_chip, help_button], spacing=8)

    page.add(
        ft.Row([
            ft.Column([
                ft.Text("DUPR Club Data Crawler", size=22, weight=ft.FontWeight.BOLD),
                ft.Text("Nhập thông tin, bấm Start để crawl và xuất Excel.", color="grey"),
            ], expand=True),
            header_right,
        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
        help_panel,
        ft.Container(content=form_row1, padding=8),
        ft.Container(content=form_row2, padding=8),
        ft.Row([start_button, stop_button, progress, status_text], spacing=16),
        ft.Row([login_button, guest_login_button, colab_button, dupr_button], spacing=16, wrap=True),
        ft.Divider(),
        ft.Text("Logs", weight=ft.FontWeight.BOLD),
        ft.Row([logs_container], alignment=ft.MainAxisAlignment.START),
        ft.Divider(),
        ft.Row([load_excel_btn], spacing=12),
        ft.Text("Công thức tính (Math)", size=16, weight=ft.FontWeight.BOLD),
        formula_plain_container,
        #formula_latex_container,
        #formula_text_container,
        formula_img1,
        formula_img2,
        mae_text,
        ft.Divider(),
        ft.Row([
            ft.Text("Biểu đồ trực quan (Visualization)", size=14, weight=ft.FontWeight.BOLD),
            ft.TextButton("?", on_click=on_toggle_visual_help),
        ], alignment=ft.MainAxisAlignment.START),
        ft.Container(
            ref=visual_help_panel,
            visible=False,
            padding=8,
            bgcolor="#FAFAFA",
            border=ft.border.all(1, "#E0E0E0"),
            border_radius=8,
            content=ft.Column([
                ft.Text("Giải thích biểu đồ", weight=ft.FontWeight.BOLD, selectable=True, color='#000000'),
                ft.Text("- Rating distribution: Phân bố rating của người chơi và đối thủ.", selectable=True, color='#000000'),
                ft.Text("- Win/Loss breakdown: Số trận thắng/thua.", selectable=True, color='#000000'),
                ft.Text("- Diff vs Delta scatter: Quan hệ giữa chênh lệch rating trước trận và mức thay đổi rating.", selectable=True, color='#000000'),
                ft.Text("- Matches over time: Số trận theo từng ngày.", selectable=True, color='#000000'),
                ft.Text("- Gender pie: Tỉ lệ Nam/Nữ dựa trên dữ liệu hồ sơ.", selectable=True, color='#000000'),
            ], spacing=4),
        ),
        ft.Row([chart_type_dd, format_filter_dd, update_chart_btn], spacing=12, wrap=True),
        chart_img,
        ft.Row([sample_dropdown], spacing=12, wrap=True),
        ft.Row([match_format, sim_player_before, sim_teammate_before, sim_opp_before, sim_opp2_before], spacing=12, wrap=True),
        ft.Row([sim_scores, sim_result, sim_example_btn, sim_compute_btn], spacing=12, wrap=True),
        result_container,
        ft.Divider(),
        ft.Container(
            alignment=ft.alignment.center,
            padding=8,
            content=ft.Row([
                ft.Text("© "),
                ft.TextButton("SportPlus", on_click=lambda _: page.launch_url("https://sportplus.vn/?fbclid=IwY2xjawMFMu1leHRuA2FlbQIxMQABHquBRE6LMRXO5M9RN-ImOWqMKoVkAosM9Dj0IGK1MzM2dOv4EcC03tvBOheA_aem_8Y2ePUrYKvR0xUQoF32odw")),
                ft.Text(" 2025  •  "),
                ft.TextButton("GitHub", on_click=lambda _: page.launch_url("https://github.com/Senju14/tool-crawl-dupr-data/tree/main")),
            ], alignment=ft.MainAxisAlignment.CENTER),
        ),
    )


if __name__ == "__main__":
    ft.app(target=main)
